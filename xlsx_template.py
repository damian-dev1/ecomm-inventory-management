import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

# File path
path = "./CompositeKey_Compare_Template.xlsx"

# Sample data generator
def generate_sample_data():
    base = [f"SKU{1000+i}" for i in range(1, 21)]
    store = [f"ST{str(i%5+1).zfill(2)}" for i in range(1, 21)]
    bin = [f"A-{i%10:02d}" for i in range(1, 21)]
    warehouse_qty = [int((i*3) % 50) for i in range(1, 21)]
    ecommerce_qty = [int((i*2) % 50) for i in range(1, 21)]

    warehouse_df = pd.DataFrame({
        "SKU": base,
        "Store": store,
        "Bin": bin,
        "Quantity": warehouse_qty
    })

    ecommerce_df = pd.DataFrame({
        "SKU": base,
        "Store": store,
        "Bin": bin,
        "Quantity": ecommerce_qty
    })

    return warehouse_df, ecommerce_df

# Add composite key and presence columns
def add_composite_key(df):
    df["CompositeKey"] = df["SKU"].astype(str).str.strip() + "|" + \
                         df["Store"].astype(str).str.strip() + "|" + \
                         df["Bin"].astype(str).str.strip()
    df["Presence"] = ""
    return df

# README content
readme_lines = [
    "**Composite-Key Compare Template**",
    "",
    "**How to use with the Tkinter app:**",
    "1. Open this workbook and replace the sample data with your own.",
    "2. Save as CSV: one for the Warehouse sheet, one for the Ecommerce sheet.",
    "3. In the app, load the CSV files into the corresponding sides.",
    "4. Build your composite key by selecting columns (e.g., SKU + Store + Bin).",
    "5. Pick the Quantity column on each side and click Compare.",
    "",
    "**Notes:**",
    "- Keep headers in row 1 exactly as-is or rename them consistently across both files.",
    "- For composite keys, you can use 1–3 columns. The app will normalize keys (trim, case, spaces).",
    "- Quantity must be numeric. Blank or non-numeric cells will be treated as 0.",
    "- Presence column in results shows whether a key exists in Warehouse, Ecommerce, or Both.",
    "",
    "**Data dictionary:**",
    "- SKU: Product identifier (string).",
    "- Store: Optional store/location code (string).",
    "- Bin: Optional bin/location within store/warehouse (string).",
    "- Quantity: On-hand or available units (integer).",
]

# Generate data
warehouse_df, ecommerce_df = generate_sample_data()
warehouse_df = add_composite_key(warehouse_df)
ecommerce_df = add_composite_key(ecommerce_df)
readme_df = pd.DataFrame({"Instructions": readme_lines})

# Write initial Excel file
with pd.ExcelWriter(path, engine="openpyxl") as writer:
    readme_df.to_excel(writer, sheet_name="README", index=False)
    warehouse_df.to_excel(writer, sheet_name="Warehouse", index=False)
    ecommerce_df.to_excel(writer, sheet_name="Ecommerce", index=False)

# Load workbook for formatting
wb = load_workbook(path)

# Styles
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="1F2937")
thin = Side(border_style="thin", color="2A2F37")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center")

# Style and format sheets
def style_sheet(ws_name):
    ws = wb[ws_name]
    ws.freeze_panes = "A2"
    widths = {"A": 16, "B": 12, "C": 12, "D": 12, "E": 30, "F": 12}
    for col_letter, width in widths.items():
        if col_letter <= get_column_letter(ws.max_column):
            ws.column_dimensions[col_letter].width = width
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = left if c != 4 else center
            cell.border = border

def add_table(ws, name):
    if ws.max_row < 2 or ws.max_column < 1:
        return
    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    table = Table(displayName=name, ref=ref)
    style = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)

for sheet in ["Warehouse", "Ecommerce"]:
    style_sheet(sheet)
    add_table(wb[sheet], f"{sheet}Table")

# README formatting
wsr = wb["README"]
wsr.column_dimensions["A"].width = 100
wsr.freeze_panes = "A2"
wsr.auto_filter.ref = None
for r in range(1, wsr.max_row + 1):
    cell = wsr.cell(row=r, column=1)
    if r == 1 or (cell.value and cell.value.startswith("**")):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
    else:
        cell.alignment = left

# Quantity validation
dv = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
dv.error = "Quantity must be a non-negative integer."
dv.errorTitle = "Invalid Quantity"
for sheet in ["Warehouse", "Ecommerce"]:
    ws = wb[sheet]
    start_row = 2
    end_row = max(1000, ws.max_row + 100)
    rng = f"D{start_row}:D{end_row}"
    ws.add_data_validation(dv)
    dv.add(rng)

# Save final file
wb.save(path)
print(f"✅ Template saved to {path}")
