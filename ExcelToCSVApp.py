import os
import re
import csv
import time
import threading
from typing import List, Dict, Optional, Tuple, Iterable
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk

import pandas as pd
import openpyxl

try:
    from ttkbootstrap import Style, ttk as ttkb  # noqa: F401
    _HAS_TTKB = True
except Exception:
    Style = None
    _HAS_TTKB = False

try:
    from pyxlsb import open_workbook as xlsb_open_workbook
    _HAS_PYXLSB = True
except Exception:
    _HAS_PYXLSB = False


def to_snake_case(header: str) -> str:
    header = re.sub(r'[^a-zA-Z0-9]+', '_', str(header))
    return header.strip('_').lower()


class CancelToken:
    def __init__(self) -> None:
        self._flag = False

    def cancel(self) -> None:
        self._flag = True

    def is_cancelled(self) -> bool:
        return self._flag


def is_xlsb(path: str | None) -> bool:
    return bool(path) and os.path.splitext(path)[1].lower() == ".xlsb"


class ExcelToCSVApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Excel → CSV Converter")
        self.root.minsize(850, 480)
        try:
            self.root.iconbitmap(default="")
        except Exception:
            pass

        if _HAS_TTKB:
            self.style = Style(theme="darkly")
        else:
            self.style = None

        self.file_path: Optional[str] = None
        self.sheet_names: List[str] = []
        self.headers: List[str] = []
        self.header_vars: Dict[str, tk.BooleanVar] = {}
        self._worker_thread: Optional[threading.Thread] = None
        self._cancel = CancelToken()

        self.pane = ttk.Panedwindow(self.root, orient="horizontal")
        self.pane.pack(fill="both", expand=True)

        self.left = ttk.Frame(self.pane, padding=10)
        self.right = ttk.Frame(self.pane, padding=10)
        self.pane.add(self.left, weight=0)
        self.pane.add(self.right, weight=3)

        self._build_left()
        self._build_right()
        self._bind_context_menus()

    def _build_left(self):
        self.left.rowconfigure(20, weight=1)
        self.left.columnconfigure(1, weight=1)

        r = 0
        ttk.Label(self.left, text="Excel → CSV Converter", font=("Segoe UI", 12, "bold")).grid(row=r, column=0, columnspan=3, sticky="w", pady=(0, 8))
        r += 1

        ttk.Button(self.left, text="Select Excel File", command=self.select_file).grid(row=r, column=0, sticky="w")
        self.file_lbl = ttk.Label(self.left, text="", foreground="#888")
        self.file_lbl.grid(row=r, column=1, columnspan=2, sticky="w", padx=8)
        r += 1

        ttk.Label(self.left, text="Sheet:").grid(row=r, column=0, sticky="w", pady=(8, 0))
        self.sheet_var = tk.StringVar()
        self.sheet_combo = ttk.Combobox(self.left, textvariable=self.sheet_var, state="readonly", width=28)
        self.sheet_combo.grid(row=r, column=1, columnspan=2, sticky="we", pady=(8, 0))
        self.sheet_combo.bind("<<ComboboxSelected>>", lambda e: self.reset_headers())
        r += 1

        ttk.Label(self.left, text="Header row:").grid(row=r, column=0, sticky="w", pady=(8, 0))
        self.header_row_var = tk.StringVar(value="1")
        self.header_row_spin = ttk.Spinbox(self.left, from_=1, to=9999, textvariable=self.header_row_var, width=6)
        self.header_row_spin.grid(row=r, column=1, sticky="w", pady=(8, 0))
        ttk.Button(self.left, text="Load Headers", command=self.load_headers).grid(row=r, column=2, sticky="e", pady=(8, 0))
        r += 1

        self.opt_snake = tk.BooleanVar(value=True)
        self.opt_dedup = tk.BooleanVar(value=True)
        ttk.Checkbutton(self.left, text="snake_case headers", variable=self.opt_snake).grid(row=r, column=0, columnspan=2, sticky="w", pady=(8, 0))
        ttk.Checkbutton(self.left, text="Remove duplicates (streaming)", variable=self.opt_dedup).grid(row=r, column=2, sticky="w", pady=(8, 0))
        r += 1

        ttk.Label(self.left, text="Columns:").grid(row=r, column=0, sticky="w", pady=(8, 2))
        r += 1

        self.columns_frame = ttk.Frame(self.left)
        self.columns_frame.grid(row=r, column=0, columnspan=3, sticky="nsew")
        self.left.rowconfigure(r, weight=1)
        self.left.columnconfigure(1, weight=1)

        self.col_canvas = tk.Canvas(self.columns_frame, borderwidth=0, highlightthickness=0)
        self.col_scroll_y = ttk.Scrollbar(self.columns_frame, orient="vertical", command=self.col_canvas.yview)
        self.col_list = ttk.Frame(self.col_canvas)
        self.col_list_id = self.col_canvas.create_window((0, 0), window=self.col_list, anchor="nw")

        self.col_canvas.configure(yscrollcommand=self.col_scroll_y.set)
        self.col_canvas.pack(side="left", fill="both", expand=True)
        self.col_scroll_y.pack(side="right", fill="y")
        self.col_list.bind("<Configure>", lambda e: self.col_canvas.configure(scrollregion=self.col_canvas.bbox("all")))
        self.col_canvas.bind("<Configure>", lambda e: self.col_canvas.itemconfigure(self.col_list_id, width=e.width))
        r += 1

        btn_row = ttk.Frame(self.left)
        btn_row.grid(row=r, column=0, columnspan=3, sticky="we", pady=(6, 0))
        ttk.Button(btn_row, text="Select All", command=lambda: self.set_all_checkboxes(True)).pack(side="left")
        ttk.Button(btn_row, text="Deselect All", command=lambda: self.set_all_checkboxes(False)).pack(side="left", padx=8)
        r += 1

        act_row = ttk.Frame(self.left)
        act_row.grid(row=r, column=0, columnspan=3, sticky="we", pady=(6, 0))
        ttk.Button(act_row, text="Preview", command=self.preview_data).pack(side="top", fill="x")
        ttk.Button(act_row, text="Export CSV (Fast)", command=self.export_csv_fast).pack(side="top", fill="x", pady=(6, 0))
        r += 1

        prog_row = ttk.Frame(self.left)
        prog_row.grid(row=r, column=0, columnspan=3, sticky="we", pady=(8, 0))
        self.prog = ttk.Progressbar(prog_row, mode="determinate", maximum=100)
        self.prog.pack(side="left", fill="x", expand=True)
        self.cancel_btn = ttk.Button(prog_row, text="Cancel", command=self.cancel_current, width=8)
        self.cancel_btn.pack(side="left", padx=(8, 0))
        r += 1

        self.status_lbl = ttk.Label(self.left, text="", foreground="#2e7d32")
        self.status_lbl.grid(row=r, column=0, columnspan=3, sticky="w", pady=(8, 0))
        r += 1

    def _build_right(self):
        right_top = ttk.Frame(self.right)
        right_top.pack(fill="x")
        ttk.Label(right_top, text="Data Preview", font=("Segoe UI", 12, "bold")).pack(side="left")
        ttk.Label(right_top, text="   Rows:").pack(side="left", padx=(10, 0))
        self.preview_rows_var = tk.IntVar(value=10)
        ttk.Spinbox(right_top, from_=1, to=2000, textvariable=self.preview_rows_var, width=6).pack(side="left", padx=(4, 0))

        tv_wrap = ttk.Frame(self.right)
        tv_wrap.pack(fill="both", expand=True, pady=(8, 6))
        self.tree = ttk.Treeview(tv_wrap, columns=(), show="headings")
        yscroll = ttk.Scrollbar(tv_wrap, orient="vertical", command=self.tree.yview)
        xscroll = ttk.Scrollbar(tv_wrap, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        yscroll.grid(row=0, column=1, sticky="ns")
        xscroll.grid(row=1, column=0, sticky="we")
        tv_wrap.rowconfigure(0, weight=1)
        tv_wrap.columnconfigure(0, weight=1)
        self.tree.bind("<Button-3>", self._tree_context_menu)

        log_frame = ttk.LabelFrame(self.right, text="Log")
        log_frame.pack(fill="x", pady=(0, 4))
        self.log_text = tk.Text(log_frame, height=5, wrap="word")
        self.log_text.pack(fill="both", expand=True)

    def _bind_context_menus(self):
        self._entry_menu = tk.Menu(self.root, tearoff=0)
        self._entry_menu.add_command(label="Paste", command=lambda: self.root.focus_get().event_generate("<<Paste>>"))

        for widget in (self.sheet_combo, self.header_row_spin): widget.bind("<Button-3>", self._show_entry_menu)
        self.root.bind_class("TEntry", "<Button-3>", self._show_entry_menu)
        self.root.bind_class("TCombobox", "<Button-3>", self._show_entry_menu)
        self.root.bind_class("TSpinbox", "<Button-3>", self._show_entry_menu)
        self.root.bind_class("Text", "<Button-3>", self._show_entry_menu)
        self.root.bind_class("Treeview", "<Button-3>", self._show_entry_menu)
        self.root.bind_class("Treeview.Heading", "<Button-3>", self._show_entry_menu)
        
        self._tree_menu = tk.Menu(self.root, tearoff=0)
        self._tree_menu.add_command(label="Copy selected row(s)", command=self._copy_selected_rows)
        self._tree_menu.add_separator()
        self._tree_menu.add_command(label="Select All", command=lambda: self.tree.selection_set(self.tree.get_children()))
        self._tree_menu.add_command(label="Deselect All", command=lambda: self.tree.selection_remove(self.tree.get_children()))
        self._tree_menu.add_separator()
        self._tree_menu.add_command(label="Clear Preview", command=self.clear_preview)
        self.tree.bind("<Button-3>", self._tree_context_menu)
        self.tree.bind("<Control-a>", lambda e: self.tree.selection_set(self.tree.get_children()))
        self.tree.bind("<Control-A>", lambda e: self.tree.selection_set(self.tree.get_children()))

    def _show_entry_menu(self, event):
        try:
            self._entry_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self._entry_menu.grab_release()

    def _tree_context_menu(self, event):
        try:
            item = self.tree.identify_row(event.y)
            if item:
                self.tree.selection_set(item)
            self._tree_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self._tree_menu.grab_release()

    def _copy_selected_rows(self):
        sel = self.tree.selection()
        if not sel:
            return
        rows = []
        cols = self.tree["columns"]
        for iid in sel:
            values = self.tree.item(iid, "values")
            row_dict = {c: v for c, v in zip(cols, values)}
            rows.append(row_dict)
        lines = ["\t".join(cols)]
        for rd in rows:
            lines.append("\t".join(str(rd.get(c, "")) for c in cols))
        txt = "\n".join(lines)
        self.root.clipboard_clear()
        self.root.clipboard_append(txt)

    # ---------------- Helpers ----------------
    def set_status(self, text: str, ok: bool = True):
        self.status_lbl.configure(text=text, foreground="#2e7d32" if ok else "#c62828")
        self._log(text)

    def _log(self, msg: str):
        self.log_text.insert("end", msg + "\n")
        self.log_text.see("end")

    def reset_headers(self):
        self.headers = []
        self.header_vars.clear()
        for w in list(self.col_list.children.values()):
            w.destroy()
        self.clear_preview()
        self.set_status("")

    def set_all_checkboxes(self, value: bool):
        for v in self.header_vars.values():
            v.set(value)

    def get_selected_columns(self) -> List[str]:
        return [c for c, v in self.header_vars.items() if v.get()]

    def clear_preview(self):
        self.tree.delete(*self.tree.get_children())
        self.tree["columns"] = ()

    def render_preview(self, df: pd.DataFrame):
        self.clear_preview()
        cols = list(df.columns)
        self.tree["columns"] = cols
        for c in cols:
            self.tree.heading(c, text=str(c))
            self.tree.column(c, width=140, stretch=True)
        for _, row in df.iterrows():
            values = [("" if pd.isna(v) else v) for v in row.tolist()]
            self.tree.insert("", "end", values=values)

    # ---------------- File / Sheet / Header ops ----------------
    def select_file(self):
        path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[
                ("Excel files", "*.xlsx *.xlsm *.xltx *.xltm *.xlsb *.xls"),
                ("All files", "*.*"),
            ]
        )
        if not path:
            return
        if is_xlsb(path) and not _HAS_PYXLSB:
            messagebox.showerror("Missing dependency",
                                 "This .xlsb file requires the 'pyxlsb' package.\n\nInstall:\n    pip install pyxlsb")
            return

        self.file_path = path
        self.file_lbl.configure(text=os.path.basename(path))
        try:
            if is_xlsb(path):
                # pyxlsb: get sheet names
                with xlsb_open_workbook(path) as wb:
                    self.sheet_names = [s.name for s in wb.sheets]
            else:
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                self.sheet_names = wb.sheetnames

            self.sheet_combo["values"] = self.sheet_names
            if self.sheet_names:
                self.sheet_var.set(self.sheet_names[0])
            self.set_status("File loaded. Pick sheet & header row, then Load Headers.", ok=True)
        except Exception as e:
            messagebox.showerror("Error", f"Could not read workbook:\n{e}")
            self.set_status("Failed to read workbook.", ok=False)

    def load_headers(self):
        if not self.file_path or not self.sheet_var.get():
            messagebox.showwarning("Missing info", "Please select a file and sheet first.")
            return
        try:
            header_row = int(self.header_row_var.get())
            if header_row < 1:
                raise ValueError
        except ValueError:
            messagebox.showerror("Invalid input", "Header row must be a positive integer.")
            return

        try:
            if is_xlsb(self.file_path):
                df = pd.read_excel(
                    self.file_path,
                    sheet_name=self.sheet_var.get(),
                    header=header_row - 1,
                    nrows=0,
                    engine="pyxlsb",
                )
            else:
                df = pd.read_excel(
                    self.file_path,
                    sheet_name=self.sheet_var.get(),
                    header=header_row - 1,
                    nrows=0,
                    engine="openpyxl",
                )

            self.headers = list(df.columns)
            self.header_vars = {h: tk.BooleanVar(value=True) for h in self.headers}

            for w in list(self.col_list.children.values()):
                w.destroy()
            for col in self.headers:
                ttk.Checkbutton(self.col_list, text=str(col), variable=self.header_vars[col]).pack(anchor="w", pady=1)

            self.set_status("Headers loaded. Select columns to preview/export.", ok=True)
        except Exception as e:
            messagebox.showerror("Error", f"Could not load headers:\n{e}")
            self.set_status("Failed to load headers.", ok=False)

    # ---------------- Preview (threaded) ----------------
    def preview_data(self):
        if not self._validate_preconditions():
            return

        def task():
            t0 = time.time()
            try:
                n = max(1, int(self.preview_rows_var.get()))
                header_row = int(self.header_row_var.get()) - 1
                selected = self.get_selected_columns()
                if is_xlsb(self.file_path):
                    df = pd.read_excel(
                        self.file_path,
                        sheet_name=self.sheet_var.get(),
                        header=header_row,
                        usecols=selected,
                        nrows=n,
                        engine="pyxlsb",
                    )
                else:
                    df = pd.read_excel(
                        self.file_path,
                        sheet_name=self.sheet_var.get(),
                        header=header_row,
                        usecols=selected,
                        nrows=n,
                        engine="openpyxl",
                    )
                self._ui(lambda: self.render_preview(df))
                self._ui(lambda: self.set_status(f"Preview loaded ({len(df)} rows) in {time.time()-t0:.2f}s.", True))
            except Exception as e:
                self._ui(lambda: messagebox.showerror("Error", f"Could not load preview:\n{e}"))
                self._ui(lambda: self.set_status("Preview failed.", False))
            finally:
                self._ui(lambda: self._progress_done())

        self._start_worker(task, indeterminate=True)

    # ---------------- Export (FAST streaming) ----------------
    def export_csv_fast(self):
        if not self._validate_preconditions():
            return

        base = os.path.splitext(os.path.basename(self.file_path))[0]
        default_name = f"{to_snake_case(base)}_{to_snake_case(self.sheet_var.get())}.csv"
        out_path = filedialog.asksaveasfilename(
            title="Save CSV As",
            defaultextension=".csv",
            initialfile=default_name,
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
        )
        if not out_path:
            return

        opt_snake = self.opt_snake.get()
        opt_dedup = self.opt_dedup.get()
        selected_cols = self.get_selected_columns()
        if not selected_cols:
            messagebox.showwarning("No columns", "Select at least one column.")
            return

        if is_xlsb(self.file_path) and not _HAS_PYXLSB:
            messagebox.showerror("Missing dependency",
                                 "This .xlsb file requires the 'pyxlsb' package.\n\nInstall:\n    pip install pyxlsb")
            return

        def task():
            t0 = time.time()
            try:
                if is_xlsb(self.file_path):
                    self._export_xlsb(out_path, selected_cols, opt_snake, opt_dedup)
                else:
                    self._export_xlsx_like(out_path, selected_cols, opt_snake, opt_dedup)

                self._ui(lambda: self.set_status(
                    f"Export completed in {time.time()-t0:.2f}s → {os.path.basename(out_path)}", True
                ))
                self._ui(lambda: messagebox.showinfo("Done", f"CSV saved to:\n{out_path}"))
            except Exception as e:
                self._ui(lambda: messagebox.showerror("Error", f"Export failed:\n{e}"))
                self._ui(lambda: self.set_status("Export failed.", False))
            finally:
                self._ui(lambda: self._progress_done())

        # xlsb progress → indeterminate (no reliable total), xlsx-like → determinate
        self._start_worker(task, indeterminate=is_xlsb(self.file_path))

    # ---- Streaming exporters ----
    def _export_xlsx_like(self, out_path: str, selected_cols: List[str], opt_snake: bool, opt_dedup: bool):
        wb = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)
        ws = wb[self.sheet_var.get()]
        header_row_idx = int(self.header_row_var.get())
        header_cells = next(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))
        header_map = {str(h): i for i, h in enumerate(header_cells)}
        missing = [c for c in selected_cols if c not in header_map]
        if missing:
            raise ValueError(f"Selected columns not found in header: {missing}")
        selected_idx = [header_map[c] for c in selected_cols]

        with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            out_header = [to_snake_case(c) if opt_snake else c for c in selected_cols]
            writer.writerow(out_header)

            seen: set = set() if opt_dedup else set()
            total_rows = ws.max_row - header_row_idx if ws.max_row and ws.max_row > header_row_idx else 0
            self._ui(lambda: self._progress_reset(total_rows))

            processed = 0
            for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
                if self._cancel.is_cancelled():
                    self._ui(lambda: self.set_status("Export cancelled by user.", False))
                    return
                out = []
                for i in selected_idx:
                    v = row[i] if i < len(row) else None
                    out.append("" if v is None else v)
                if opt_dedup:
                    key = tuple(out)
                    if key in seen:
                        processed += 1
                        if total_rows and processed % 100 == 0:
                            self._ui(lambda: self._progress_step(100))
                        continue
                    seen.add(key)
                writer.writerow(out)
                processed += 1
                if total_rows and (processed % 100 == 0 or processed == total_rows):
                    self._ui(lambda: self._progress_set(min(processed, total_rows)))

    def _export_xlsb(self, out_path: str, selected_cols: List[str], opt_snake: bool, opt_dedup: bool):
        # Indeterminate progress; stream rows with pyxlsb
        with xlsb_open_workbook(self.file_path) as wb:
            sheet = wb.get_sheet(self.sheet_var.get())
            header_row_idx = int(self.header_row_var.get())

            # Build header from the specified row (1-based)
            hdr_vals: List[str] = []
            for idx, row in enumerate(sheet.rows(), start=1):
                if idx == header_row_idx:
                    hdr_vals = [None if c is None else str(c.v) if hasattr(c, "v") else str(c) for c in row]
                    break
            if not hdr_vals:
                raise ValueError("Header row not found in .xlsb sheet.")
            header_map = {str(h): i for i, h in enumerate(hdr_vals)}
            missing = [c for c in selected_cols if c not in header_map]
            if missing:
                raise ValueError(f"Selected columns not found in header: {missing}")
            selected_idx = [header_map[c] for c in selected_cols]

            with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                out_header = [to_snake_case(c) if opt_snake else c for c in selected_cols]
                writer.writerow(out_header)

                seen: set = set() if opt_dedup else set()
                self._ui(lambda: self.prog.configure(mode="indeterminate"))
                self._ui(lambda: self.prog.start(10))

                # Continue streaming rows after header
                for idx, row in enumerate(sheet.rows(), start=1):
                    if idx <= header_row_idx:
                        continue
                    if self._cancel.is_cancelled():
                        self._ui(lambda: self.set_status("Export cancelled by user.", False))
                        return
                    vals = []
                    for i in selected_idx:
                        v = row[i].v if i < len(row) and hasattr(row[i], "v") else (row[i] if i < len(row) else None)
                        vals.append("" if v is None else v)
                    if opt_dedup:
                        key = tuple(vals)
                        if key in seen:
                            continue
                        seen.add(key)
                    writer.writerow(vals)

    # ---------------- Worker / Progress ----------------
    def _start_worker(self, target, indeterminate: bool):
        if self._worker_thread and self._worker_thread.is_alive():
            messagebox.showwarning("Busy", "Please wait for the current task to finish or cancel it.")
            return
        self._cancel = CancelToken()
        self._progress_reset(0 if not indeterminate else None)
        if indeterminate:
            self.prog.configure(mode="indeterminate")
            self.prog.start(10)
        else:
            self.prog.configure(mode="determinate")
        self.cancel_btn.configure(state="normal")

        self._worker_thread = threading.Thread(target=target, daemon=True)
        self._worker_thread.start()

    def cancel_current(self):
        self._cancel.cancel()

    def _progress_reset(self, maximum: Optional[int]):
        if maximum is None:
            return
        self.prog.configure(mode="determinate", maximum=max(1, int(maximum)))
        self.prog["value"] = 0

    def _progress_set(self, value: int):
        try:
            self.prog["value"] = value
        except Exception:
            pass

    def _progress_step(self, step: int):
        try:
            self.prog.step(step)
        except Exception:
            pass

    def _progress_done(self):
        try:
            if str(self.prog["mode"]) == "indeterminate":
                self.prog.stop()
            self.prog["value"] = self.prog["maximum"] if self.prog["mode"] == "determinate" else 0
            self.cancel_btn.configure(state="disabled")
        except Exception:
            pass

    def _ui(self, fn):
        self.root.after(0, fn)

    def _validate_preconditions(self) -> bool:
        if not self.file_path or not self.sheet_var.get():
            messagebox.showwarning("Missing info", "Please select a file and sheet first.")
            return False
        if not self.headers:
            messagebox.showwarning("No headers", "Click 'Load Headers' first.")
            return False
        if not self.get_selected_columns():
            messagebox.showwarning("No columns", "Select at least one column.")
            return False
        if is_xlsb(self.file_path) and not _HAS_PYXLSB:
            messagebox.showerror("Missing dependency",
                                 "This .xlsb file requires the 'pyxlsb' package.\n\nInstall:\n    pip install pyxlsb")
            return False
        return True


if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelToCSVApp(root)
    root.mainloop()
